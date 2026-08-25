"""Scraper CONOSCE — Portal de Datos Abiertos del OSCE (Pentaho BI).

Fuentes:
1. Convocatorias: bi.seace.gob.pe/pentaho portal con links a XLSX por anio
   -> Redirect a conosce.osce.gob.pe/buscador/assets/.../convocatorias/{year}/
2. PAC (Plan Anual de Contrataciones): mismo portal, seccion PAC
   -> Redirect a conosce.osce.gob.pe/buscador/assets/.../pac/{year}/
3. Contratos: mismo portal, seccion contratos

Estrategia:
1. Scrape la pagina HTML del portal Pentaho para obtener los tinyurl por anio
2. Resolve los tinyurl para obtener las URLs directas al XLSX
3. Descarga y parsea el XLSX del anio actual
4. Filtra y upsert nuevas licitaciones

Nota: El scraper ocds_api.py ya descarga convocatorias. Este scraper
se enfoca en los CONTRATOS (adjudicados) y el PAC (planeado) como
fuentes complementarias.
"""
import logging
import hashlib
import os
import io
import re
from datetime import datetime, date
from pathlib import Path

import httpx
import openpyxl
from bs4 import BeautifulSoup
from tenacity import retry, stop_after_attempt, wait_exponential

from shared.db import upsert_licitacion, log_scraping_start, log_scraping_end, get_config

log = logging.getLogger("radar.conosce")

# Pentaho portal pages
PORTAL_CONTRATOS = (
    "https://bi.seace.gob.pe/pentaho/api/repos/"
    ":public:portal:datosabiertoscontratos.html/content"
)
PORTAL_PAC = (
    "https://bi.seace.gob.pe/pentaho/api/repos/"
    ":public:portal:datosabiertospac.html/content"
)

# Direct URL patterns (fallback if portal scraping fails)
CONTRATOS_DIRECT_URL = (
    "https://conosce.osce.gob.pe/buscador/assets/67ae6c4a"
    "/reportes/contratos/{year}/CONOSCE_CONTRATOS{year}_0.xlsx"
)
PAC_DIRECT_URL = (
    "https://conosce.osce.gob.pe/buscador/assets/67ae6c4a"
    "/reportes/pac/{year}/CONOSCE_PAC{year}_0.xlsx"
)

PENTAHO_PARAMS = {"userid": "public", "password": "key"}

CACHE_DIR = Path(os.getenv("CACHE_DIR", "/tmp/licitapro_cache"))

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
}

# ---------- Columnas CONTRATOS XLSX ----------
# Verificado contra la estructura real del archivo:
# 0=codigoentidad, 1=codigoconvocatoria, 2=descripcion_proceso,
# 3=n_cod_contrato, 4=codigo_contrato, 5=num_contrato, 6=num_item,
# 7=monto_contratado_total, 8=monto_contratado_item, ...
# 13=moneda, 17=urlcontrato, 19=fecha_publicacion_contrato
CONT_COL_CODIGO_ENTIDAD = 0
CONT_COL_CODIGO_CONV = 1
CONT_COL_DESC_PROCESO = 2
CONT_COL_NUM_CONTRATO = 5
CONT_COL_MONTO_TOTAL = 7
CONT_COL_MONEDA = 13
CONT_COL_URL_CONTRATO = 17
CONT_COL_FECHA_PUB = 19

# ---------- Columnas PAC XLSX ----------
PAC_COL_ENTIDAD = 3
PAC_COL_DESC_PROCESO = 6
PAC_COL_TIPO_PROC = 7
PAC_COL_OBJETO = 8
PAC_COL_DESC_ITEM = 10
PAC_COL_DEPTO = 13
PAC_COL_PROVINCIA = 14

# Mapeo de departamentos
DEPTOS_MAP = {
    "MADRE DE DIOS": "Madre de Dios", "CUSCO": "Cusco", "JUNIN": "Junín",
    "JUNÍN": "Junín", "LIMA": "Lima", "AREQUIPA": "Arequipa", "PUNO": "Puno",
    "PIURA": "Piura", "LA LIBERTAD": "La Libertad", "LAMBAYEQUE": "Lambayeque",
    "LORETO": "Loreto", "UCAYALI": "Ucayali", "ANCASH": "Áncash",
    "ÁNCASH": "Áncash", "CAJAMARCA": "Cajamarca", "HUANUCO": "Huánuco",
    "HUÁNUCO": "Huánuco", "ICA": "Ica", "TACNA": "Tacna",
    "MOQUEGUA": "Moquegua", "TUMBES": "Tumbes", "AMAZONAS": "Amazonas",
    "PASCO": "Pasco", "AYACUCHO": "Ayacucho", "APURIMAC": "Apurímac",
    "APURÍMAC": "Apurímac", "HUANCAVELICA": "Huancavelica",
    "SAN MARTIN": "San Martín", "SAN MARTÍN": "San Martín",
    "CALLAO": "Callao",
}


def _normalize_depto(raw: str | None) -> str | None:
    if not raw:
        return None
    return DEPTOS_MAP.get(raw.strip().upper())


def _parse_monto(val) -> float | None:
    if val is None:
        return None
    try:
        v = float(val)
        return v if v > 0 else None
    except (ValueError, TypeError):
        return None


def _parse_fecha(val) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime.combine(val, datetime.min.time())
    text = str(val).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%Y %H:%M", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _detect_tipo(nomenclatura: str) -> str:
    n = (nomenclatura or "").upper()
    if "LP-" in n or "LICITACION" in n:
        return "LP"
    if "AS-" in n or "ADJUDICACION" in n:
        return "AS"
    if "SIE-" in n or "SUBASTA" in n:
        return "SIE"
    if "CP-" in n or "CONCURSO" in n:
        return "CP"
    if "CD-" in n or "CONTRATACION DIRECTA" in n:
        return "CD"
    if "COMPARACION" in n or "CDP" in n or "CATALOGO" in n:
        return "CdP"
    return "otro"


def _gen_id(prefix: str, *parts) -> str:
    raw = f"{prefix}_{'_'.join(str(p) for p in parts)}".lower().strip()
    return hashlib.md5(raw.encode()).hexdigest()[:16]


def _match_keywords(texto: str, keywords: list[str]) -> bool:
    # Delega en el helper compartido, que ignora tildes: las fuentes de OSCE
    # traen la acentuacion corrupta y el match con tilde nunca ocurria.
    from shared.config import match_keywords
    return match_keywords(texto, keywords)


async def _resolve_tinyurl(client: httpx.AsyncClient, tinyurl: str) -> str | None:
    """Resolve un tinyurl a la URL final."""
    try:
        r = await client.head(tinyurl, follow_redirects=True)
        if r.status_code == 200:
            return str(r.url)
    except Exception:
        pass
    # Fallback: GET con follow
    try:
        r = await client.get(tinyurl, follow_redirects=False)
        if r.status_code in (301, 302, 303, 307):
            return r.headers.get("location")
    except Exception:
        pass
    return None


async def _get_portal_links(client: httpx.AsyncClient, portal_url: str) -> dict[int, str]:
    """Scrape un portal Pentaho y extrae los tinyurl por anio.

    Retorna {year: resolved_url}.
    """
    try:
        resp = await client.get(portal_url, params=PENTAHO_PARAMS)
        if resp.status_code != 200:
            log.warning(f"CONOSCE portal returned {resp.status_code}")
            return {}
    except Exception as e:
        log.error(f"CONOSCE portal fetch failed: {e}")
        return {}

    soup = BeautifulSoup(resp.text, "html.parser")
    year_links: dict[int, str] = {}

    # El portal tiene secciones h2 con el anio, y links tinyurl
    # Buscar todos los links con tinyurl
    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"]
        text = a_tag.get_text(strip=True)

        if "tinyurl.com" not in href:
            continue
        if "TODOS LOS PROCESOS" not in text.upper() and "DESCARGAR" not in text.upper():
            # Solo nos interesan los links de "DESCARGAR TODOS LOS PROCESOS" o "DESCARGAR"
            if "SIE" in text.upper() or "DIRECTA" in text.upper():
                continue

        # Extraer el anio del tinyurl (patron: conosceconvocatorias2025)
        year_match = re.search(r"(\d{4})", href)
        if not year_match:
            # Buscar el anio en el contexto cercano
            parent = a_tag.find_parent("div")
            if parent:
                h2 = parent.find_previous("h2")
                if h2:
                    year_match = re.search(r"(\d{4})", h2.get_text())

        if year_match:
            year = int(year_match.group(1))
            if 2018 <= year <= 2030:
                year_links[year] = href

    return year_links


async def _download_xlsx_cached(
    client: httpx.AsyncClient, url: str, cache_key: str
) -> bytes | None:
    """Descarga un XLSX con cache local."""
    cache_file = CACHE_DIR / f"{cache_key}.xlsx"
    cache_size_file = CACHE_DIR / f"{cache_key}.size"

    # Check remote size
    remote_size = 0
    try:
        head = await client.head(url, follow_redirects=True)
        if head.status_code != 200:
            return None
        remote_size = int(head.headers.get("content-length", 0))
    except Exception:
        pass

    # Check cache
    if cache_file.exists() and cache_size_file.exists():
        try:
            cached_size = int(cache_size_file.read_text().strip())
            if remote_size > 0 and cached_size == remote_size:
                log.info(f"CONOSCE: Using cached {cache_key} ({remote_size} bytes)")
                return cache_file.read_bytes()
        except (ValueError, OSError):
            pass

    # Download
    log.info(f"CONOSCE: Downloading {cache_key} from {url}")
    try:
        resp = await client.get(url, follow_redirects=True)
        if resp.status_code != 200:
            return None
        data = resp.content
        log.info(f"CONOSCE: Downloaded {len(data)} bytes")
    except Exception as e:
        log.error(f"CONOSCE: Download failed: {e}")
        return None

    # Cache
    try:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cache_file.write_bytes(data)
        cache_size_file.write_text(str(len(data)))
    except OSError:
        pass

    return data


def _detect_depto_from_text(texto: str) -> str | None:
    """Intenta detectar departamento desde texto libre."""
    upper = texto.upper()
    for key, val in DEPTOS_MAP.items():
        if key in upper:
            return val
    return None


def _parse_contratos_xlsx(xlsx_data: bytes, config: dict) -> list[dict]:
    """Parsea el XLSX de contratos.

    El XLSX de contratos tiene columnas: codigoentidad, codigoconvocatoria,
    descripcion_proceso, num_contrato, monto_contratado_total, urlcontrato, etc.
    No tiene columna de nombre de entidad ni departamento, asi que
    detectamos departamento desde la descripcion del proceso.
    """
    keywords = config.get("keywords", [])
    regiones = config.get("regiones", [])
    monto_min = config.get("monto_min", 0) or 0
    monto_max = config.get("monto_max", 999999999) or 999999999

    results = []
    seen = set()

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_data), read_only=True)
    ws = wb.active

    row_num = 0
    for row in ws.iter_rows(values_only=True):
        row_num += 1
        if row_num == 1:
            continue

        cols = list(row) + [None] * 30

        codigo_entidad = str(cols[CONT_COL_CODIGO_ENTIDAD] or "").strip()
        codigo_conv = str(cols[CONT_COL_CODIGO_CONV] or "").strip()
        desc = str(cols[CONT_COL_DESC_PROCESO] or "").strip()
        num_contrato = str(cols[CONT_COL_NUM_CONTRATO] or "").strip()
        monto_raw = cols[CONT_COL_MONTO_TOTAL]
        url_contrato = str(cols[CONT_COL_URL_CONTRATO] or "").strip()
        fecha_raw = cols[CONT_COL_FECHA_PUB]

        # Deduplicar por codigo_conv + num_contrato
        dedup_key = f"{codigo_conv}_{num_contrato}"
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        if not desc or len(desc) < 10:
            continue

        monto = _parse_monto(monto_raw)

        # Detectar departamento desde la descripcion del proceso
        depto = _detect_depto_from_text(desc)

        # Filtros
        if regiones and depto and depto not in regiones:
            continue
        if monto and (monto < monto_min or monto > monto_max):
            continue
        if keywords and not _match_keywords(desc, keywords):
            continue

        # URL: usar la de SEACE si existe, sino construir una de CONOSCE
        final_url = url_contrato if url_contrato.startswith("http") else \
            f"https://conosce.osce.gob.pe/buscador/detalle?convocatoria={codigo_conv}"

        results.append({
            "id": _gen_id("contrato", codigo_conv, num_contrato),
            "fuente": "conosce_contratos",
            "tipo": _detect_tipo(num_contrato + " " + desc),
            "nomenclatura": num_contrato,
            "entidad": f"Entidad {codigo_entidad}",  # No hay nombre en este XLSX
            "entidad_tipo": None,
            "objeto": desc,
            "descripcion": None,
            "monto_referencial": monto,
            "moneda": "PEN",
            "departamento": depto,
            "url": final_url,
            "estado": "adjudicado",
            "fecha_publicacion": _parse_fecha(fecha_raw),
            "fecha_cierre": None,
        })

    wb.close()
    return results


def _parse_pac_xlsx(xlsx_data: bytes, config: dict) -> list[dict]:
    """Parsea el XLSX del PAC (Plan Anual de Contrataciones)."""
    keywords = config.get("keywords", [])
    regiones = config.get("regiones", [])
    monto_min = config.get("monto_min", 0) or 0
    monto_max = config.get("monto_max", 999999999) or 999999999

    results = []
    seen = set()

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_data), read_only=True)
    ws = wb.active

    row_num = 0
    for row in ws.iter_rows(values_only=True):
        row_num += 1
        if row_num == 1:
            continue

        cols = list(row) + [None] * 25

        entidad = str(cols[PAC_COL_ENTIDAD] or "").strip()
        desc_proceso = str(cols[PAC_COL_DESC_PROCESO] or "").strip()
        tipo_proc = str(cols[PAC_COL_TIPO_PROC] or "").strip()
        objeto = str(cols[PAC_COL_OBJETO] or "").strip()
        desc_item = str(cols[PAC_COL_DESC_ITEM] or "").strip()
        depto_raw = str(cols[PAC_COL_DEPTO] or "").strip() if cols[PAC_COL_DEPTO] else None

        # Deduplicar por entidad+descripcion
        dedup_key = f"{entidad}_{desc_proceso[:50]}"
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        if not entidad or not (desc_proceso or desc_item):
            continue

        depto = _normalize_depto(depto_raw)

        # Filtros
        if regiones and depto and depto not in regiones:
            continue
        texto = f"{desc_proceso} {desc_item} {entidad} {objeto}"
        if keywords and not _match_keywords(texto, keywords):
            continue

        # PAC no tiene monto individual por default,
        # pero podemos intentar parsear si existe
        results.append({
            "id": _gen_id("pac", entidad[:30], desc_proceso[:40]),
            "fuente": "conosce_pac",
            "tipo": _detect_tipo(tipo_proc),
            "nomenclatura": "",
            "entidad": entidad,
            "entidad_tipo": None,
            "objeto": desc_proceso or desc_item,
            "descripcion": desc_item,
            "monto_referencial": None,
            "moneda": "PEN",
            "departamento": depto,
            "url": "https://conosce.osce.gob.pe/buscador/",
            "estado": "planificado",
            "fecha_publicacion": None,
            "fecha_cierre": None,
        })

    wb.close()
    return results


@retry(stop=stop_after_attempt(2), wait=wait_exponential(min=5, max=30))
async def scrape_conosce(user_id: int = 0) -> list[dict]:
    """Scrape CONOSCE: contratos adjudicados y PAC del anio actual."""
    log_id = await log_scraping_start("conosce")
    config = await get_config(user_id)
    if not config:
        config = {}

    nuevas = []
    encontradas = 0
    errores = 0
    error_detalle = None

    current_year = date.today().year

    try:
        async with httpx.AsyncClient(
            timeout=120, headers=HEADERS, follow_redirects=True
        ) as client:

            # --- 1. Contratos del anio actual ---
            try:
                contratos_url = CONTRATOS_DIRECT_URL.format(year=current_year)
                xlsx_data = await _download_xlsx_cached(
                    client, contratos_url, f"contratos_{current_year}"
                )
                if xlsx_data:
                    parsed = _parse_contratos_xlsx(xlsx_data, config)
                    encontradas += len(parsed)
                    log.info(f"CONOSCE contratos {current_year}: {len(parsed)} matched")

                    for lic in parsed:
                        try:
                            is_new = await upsert_licitacion(lic)
                            if is_new:
                                nuevas.append(lic)
                        except Exception as e:
                            errores += 1
                            log.error(f"CONOSCE upsert error: {e}")
                else:
                    log.warning(f"CONOSCE: No contratos data for {current_year}")
            except Exception as e:
                errores += 1
                log.error(f"CONOSCE contratos error: {e}")

            # --- 2. PAC del anio actual ---
            try:
                pac_url = PAC_DIRECT_URL.format(year=current_year)
                xlsx_data = await _download_xlsx_cached(
                    client, pac_url, f"pac_{current_year}"
                )
                if xlsx_data:
                    parsed = _parse_pac_xlsx(xlsx_data, config)
                    encontradas += len(parsed)
                    log.info(f"CONOSCE PAC {current_year}: {len(parsed)} matched")

                    for lic in parsed:
                        try:
                            is_new = await upsert_licitacion(lic)
                            if is_new:
                                nuevas.append(lic)
                        except Exception as e:
                            errores += 1
                            log.error(f"CONOSCE PAC upsert error: {e}")
                else:
                    log.warning(f"CONOSCE: No PAC data for {current_year}")
            except Exception as e:
                errores += 1
                log.error(f"CONOSCE PAC error: {e}")

    except Exception as e:
        errores += 1
        error_detalle = str(e)[:200]
        log.error(f"CONOSCE scraping failed: {e}")

    await log_scraping_end(log_id, encontradas, len(nuevas), errores, error_detalle)
    log.info(
        f"CONOSCE: {encontradas} encontradas, {len(nuevas)} nuevas, {errores} errores"
    )
    return nuevas
